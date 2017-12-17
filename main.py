from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import datetime
import random
import os.path

APP_INFO = "TMxVeri Spreadsheet Creator v1.0"
APP_DESCR = APP_INFO + "\n\n17 Nov 2017\n\nUsed to populate the fixed format MS Excel template\n" \
                       "with the dummy data of 5 consecutive Serveron TM8 verifications.\n\nSource code is available " \
                       "on GitHub"
TEMPLATE_NAME = "template.xlsx"


class Gas:

    def __init__(self, n, f, b_rt=0.0, d_rt=0.0, b_ps=0.0, d_ps=0.0, rt=0.0, ps=0.0):
        self.name = n
        self.formula = f
        self.base_retention_time = b_rt
        self.distr_retention_time = d_rt
        self.base_peak_square = b_ps
        self.distr_peak_square = d_ps
        self.retention_time = rt
        self.peak_square = ps
        self.__status = ""

    def __set_status(self):
        self.__status = "Gas " + self.name + " (" + self.formula + "), rt = " + str(self.retention_time) + ", ps = " \
                        + str(self.peak_square)

    def get_status(self):
        self.__set_status()
        return self.__status

    def generate_retention_time(self):
        self.retention_time = random.uniform(self.base_retention_time - self.distr_retention_time,
                                               self.base_retention_time + self.distr_retention_time)
        self.__set_status()

    def generate_peak_square(self):
        self.peak_square = random.uniform(self.base_peak_square - self.distr_peak_square,
                                            self.base_peak_square + self.distr_peak_square)
        self.__set_status()

    def generate_all(self):
        self.generate_retention_time()
        self.generate_peak_square()


class VerificationData:

    def __init__(self):
        self.date_time = datetime.datetime.now().strftime("%d.%m.%Y")
        self.number = 0
        self.temperature = 0
        self.tcd_noise = 0
        self.gases = [Gas("Acetylene", "C2H2", 281.0, 2.0, 5900.0, 100.0),
                      Gas("Ethylene", "C2H4", 250.0, 2.0, 18510.0, 200.0),
                      Gas("Methane", "CH4", 807.0, 3.0, 9305.0, 100.0),
                      Gas("Ethane", "C2H6", 301.5, 2.0, 13.8, 0.1),
                      Gas("Carbon Dioxide", "CO2", 159.0, 1.0, 176000.0, 1000.0),
                      Gas("Oxygen", "O2", 545.5, 3.0, 1840000.0, 10000.0),
                      Gas("Carbon Oxide", "CO", 978.0, 3.0, 159.0, 1.0),
                      Gas("Hydrogen", "H2", 447.0, 2.0, 100.0, 1.0)]
        self.__status = ""

    def get_status(self):
        s = ""
        for i in self.gases: s += i.get_status() + "\n"
        self.__status = "Chromatogram #" + str(self.number) + \
               "\nDate " + str(self.date_time) + \
               "\nSample temperature " + str(self.temperature) + \
               "\nTCD noise " + str(self.tcd_noise) + "\n" + s + "\n"
        return self.__status


class ChromatographData:

    def __init__(self, sn=""):
        self.serial_no = sn
        self.verifications = [VerificationData() for i in range(0, 5)]
        self.__status = ""

    def __set_status(self):
        s = "Chromatograph, serial no. " + self.serial_no + "\n\n"
        for i in self.verifications: s += i.get_status()
        self.__status = s

    def get_status(self):
        # self.__set_status()
        return self.__status

    def generate_chromatogram_date_time(self):
        for i in self.verifications:
            i.date_time = datetime.datetime.now().strftime("%d.%m.%Y")

    def generate_chromatogram_number(self):
        b = random.randrange(1, 11)
        k = 0
        for i in self.verifications:
            i.number = b + k
            k += 1

    def generate_chromatogram_temperature(self):
        for i in self.verifications:
            i.temperature = random.randrange(3150, 3170) / 100

    def generate_chromatogram_tcd_noise(self):
        for i in self.verifications:
            i.tcd_noise = random.randrange(40,90) / 100

    def generate_chromatogram_all(self):
        self.generate_chromatogram_date_time()
        self.generate_chromatogram_number()
        self.generate_chromatogram_temperature()
        self.generate_chromatogram_tcd_noise()
        for i in self.verifications:
            for k in i.gases:
                k.generate_all()
        self.__set_status()


class ExcelHandler:

    def __init__(self):
        if os.path.isfile(TEMPLATE_NAME):
            self.work_book = load_workbook(TEMPLATE_NAME)
            self.__status = "File " + TEMPLATE_NAME + " is found and loaded.\n"

            # Excel worksheet cells
            self.cell_chromatograph_serial_number = 'C1'

            data_range6 = range(6, 11)
            self.cell_verification_date = ['B' + str(i) for i in data_range6]
            self.cell_probe_temperature = ['C' + str(i) for i in data_range6]
            self.cell_chromatogram_number = ['D' + str(i) for i in data_range6]
            self.cell_chromatogram_tcd_noise = ['E' + str(i) for i in data_range6]
            self.cell_chromatogram_c2h2_retention_time = ['F' + str(i) for i in data_range6]
            self.cell_chromatogram_c2h2_peak_square = ['G' + str(i) for i in data_range6]
            self.cell_chromatogram_c2h4_retention_time = ['H' + str(i) for i in data_range6]
            self.cell_chromatogram_c2h4_peak_square = ['I' + str(i) for i in data_range6]
            self.cell_chromatogram_ch4_retention_time = ['J' + str(i) for i in data_range6]
            self.cell_chromatogram_ch4_peak_square = ['K' + str(i) for i in data_range6]
            self.cell_chromatogram_c2h6_retention_time = ['L' + str(i) for i in data_range6]
            self.cell_chromatogram_c2h6_peak_square = ['M' + str(i) for i in data_range6]
            self.cell_chromatogram_co2_retention_time = ['N' + str(i) for i in data_range6]
            self.cell_chromatogram_co2_peak_square = ['O' + str(i) for i in data_range6]
            self.cell_chromatogram_o2_retention_time = ['P' + str(i) for i in data_range6]
            self.cell_chromatogram_o2_peak_square = ['Q' + str(i) for i in data_range6]
            self.cell_chromatogram_co_retention_time = ['R' + str(i) for i in data_range6]
            self.cell_chromatogram_co_peak_square = ['S' + str(i) for i in data_range6]
            self.cell_chromatogram_h2_retention_time = ['T' + str(i) for i in data_range6]
            self.cell_chromatogram_h2_peak_square = ['U' + str(i) for i in data_range6]
        else:
            self.__status = "File " + TEMPLATE_NAME + " is not found, the data is not loaded.\n"

    def get_status(self):
        return self.__status

    def save_file(self, cp):
        try:
            ws = self.work_book.active
            ws.title = cp.serial_no

            for i in range(0, 5):
                ws[self.cell_verification_date[i]] = cp.verifications[i].date_time
                ws[self.cell_probe_temperature[i]] = cp.verifications[i].temperature
                ws[self.cell_chromatogram_number[i]] = cp.verifications[i].number
                ws[self.cell_chromatogram_tcd_noise[i]] = cp.verifications[i].tcd_noise
                ws[self.cell_chromatogram_c2h2_retention_time[i]] = cp.verifications[i].gases[0].retention_time
                ws[self.cell_chromatogram_c2h2_peak_square[i]] = cp.verifications[i].gases[0].peak_square
                ws[self.cell_chromatogram_c2h4_retention_time[i]] = cp.verifications[i].gases[1].retention_time
                ws[self.cell_chromatogram_c2h4_peak_square[i]] = cp.verifications[i].gases[1].peak_square
                ws[self.cell_chromatogram_ch4_retention_time[i]] = cp.verifications[i].gases[2].retention_time
                ws[self.cell_chromatogram_ch4_peak_square[i]] = cp.verifications[i].gases[2].peak_square
                ws[self.cell_chromatogram_c2h6_retention_time[i]] = cp.verifications[i].gases[3].retention_time
                ws[self.cell_chromatogram_c2h6_peak_square[i]] = cp.verifications[i].gases[3].peak_square
                ws[self.cell_chromatogram_co2_retention_time[i]] = cp.verifications[i].gases[4].retention_time
                ws[self.cell_chromatogram_co2_peak_square[i]] = cp.verifications[i].gases[4].peak_square
                ws[self.cell_chromatogram_o2_retention_time[i]] = cp.verifications[i].gases[5].retention_time
                ws[self.cell_chromatogram_o2_peak_square[i]] = cp.verifications[i].gases[5].peak_square
                ws[self.cell_chromatogram_co_retention_time[i]] = cp.verifications[i].gases[6].retention_time
                ws[self.cell_chromatogram_co_peak_square[i]] = cp.verifications[i].gases[6].peak_square
                ws[self.cell_chromatogram_h2_retention_time[i]] = cp.verifications[i].gases[7].retention_time
                ws[self.cell_chromatogram_h2_peak_square[i]] = cp.verifications[i].gases[7].peak_square

            self.work_book.save(cp.serial_no + ".xlsx")
            self.__status = "Saved succesfully to" + cp.serial_no + ".xlsx.\n"
        except Exception as e:
            self.__status = "Something went wrong, the file is not saved.\n>>> " + str(e) + "\n"


class Application:

    def __init__(self):

        # chromatograph data
        self.chromatograph_data = ChromatographData()

        # application interface
        self.root = Tk()
        self.root.title(APP_INFO)

        self.menubar = Menu(self.root)

        self.menu_file = Menu(self.menubar, tearoff=0)
        self.menu_file.add_command(label="Exit", command=self.root.quit)

        self.menu_edit = Menu(self.menubar, tearoff=0)
        self.menu_edit.add_command(label="Clear text", command=self.clear_text)
        self.menu_edit.add_command(label="Forget all data", command=self.forget_all_data)

        self.menu_info = Menu(self.menubar, tearoff=0)
        self.menu_info.add_command(label="About", command=self.show_about_window)

        self.menubar.add_cascade(label="File", menu=self.menu_file)
        self.menubar.add_cascade(label="Edit", menu=self.menu_edit)
        self.menubar.add_cascade(label="Info", menu=self.menu_info)
        self.root.config(menu=self.menubar)

        self.frame_left = ttk.Frame(self.root, relief=RAISED)
        self.frame_right = ttk.Frame(self.root, relief=RAISED)
        self.root.columnconfigure(0, weight=0)
        self.root.columnconfigure(1, weight=5)
        self.root.rowconfigure(0, weight=1)

        self.frame_left.grid(row=0, column=0, sticky='snew')
        self.frame_right.grid(row=0, column=1, sticky='snew')
        self.frame_right.rowconfigure(0, weight=1)
        self.frame_right.columnconfigure(0, weight=1)
        self.frame_right.rowconfigure(1, weight=0)
        self.frame_right.columnconfigure(1, weight=0)

        self.button_generate_data = ttk.Button(self.frame_left, text="Generate data",
                                               command=self.generate_data)
        self.button_generate_data.grid(row=0, column=0, padx=25, pady=15, sticky='snew')
        self.button_save_spreadsheet = ttk.Button(self.frame_left, text="Save spreadsheet",
                                                  command=self.save_spreadsheet)
        self.button_save_spreadsheet.grid(row=1, column=0, padx=25, pady=5, sticky='snew')
        self.label_tmx_serial_no = ttk.Label(self.frame_left, text="TMx serial no.:")
        self.label_tmx_serial_no.grid(row=2, column=0)
        self.entry_tmx_serial_no = ttk.Entry(self.frame_left)
        self.entry_tmx_serial_no.grid(row=3, column=0)

        self.text_data = Text(self.frame_right)
        self.text_data.grid(row=0, column=0, sticky='snew')
        self.y_scrollbar = ttk.Scrollbar(self.frame_right, orient=VERTICAL, command=self.text_data.yview)
        self.y_scrollbar.grid(row=0, column=1, sticky='sn')
        self.x_scrollbar = ttk.Scrollbar(self.frame_right, orient=HORIZONTAL, command=self.text_data.xview)
        self.x_scrollbar.grid(row=1, column=0, sticky='ew')
        self.text_data.configure(xscrollcommand=self.x_scrollbar.set, yscrollcommand=self.y_scrollbar.set)

        self.add_status_string(APP_INFO)

        self.root.mainloop()

    def add_status_string(self, s):
        self.text_data.insert(END, ">>> " + s + "\n")
        self.text_data.see(END)

    def generate_data(self):
        self.chromatograph_data.serial_no = self.entry_tmx_serial_no.get()
        self.chromatograph_data.generate_chromatogram_all()
        self.add_status_string("========================================\n")
        self.add_status_string("Data is generated\n")
        self.add_status_string(self.chromatograph_data.get_status())

    def save_spreadsheet(self):
        # print(self.chromatograph_data.get_status())

        if self.chromatograph_data.get_status() == "":
            self.add_status_string("Generate data first.")
        else:
            self.excel_handler = ExcelHandler()
            self.add_status_string(self.excel_handler.get_status())
            self.excel_handler.save_file(self.chromatograph_data)
            self.add_status_string(self.excel_handler.get_status())

    def clear_text(self):
        self.text_data.delete('1.0', END)
        self.add_status_string(APP_INFO)

    def forget_all_data(self):
        self.chromatograph_data = ChromatographData()
        self.entry_tmx_serial_no.delete(0, END)
        self.clear_text()

    def show_about_window(self):
        messagebox.showinfo(APP_INFO, APP_DESCR)


if __name__ == "__main__": Application()
